print ('hello')
import pandas as pd
import numpy as np
import openpyxl
# from copy import deepcopy
from docx import Document
# document = Document('chinese_story.docx')
import docx2txt




# Import the story and put it into a list

story = docx2txt.process('chinese_story.docx')
story_list = []
for i in story:
    story_list.append(i)


# remove the grammar

print ('the story is currently ', len(story_list), 'words long')

for i in story:
    if i is '\n' or i is '.' or i is ',' or i is '。' or i is ' ': # remove all grammar stuff
        #including removing spaces and chinese full stops --> 。
        story_list.remove(i)
    if i is '。':
        story_list.remove(i)


print ('the story is now ', len(story_list), 'words long after removing grammar')



# work around, import the vocab into a list from word document

vocab_from_word = docx2txt.process('chinese_vocab_in_a_list.docx')
vocab_from_word_in_list = []
for i in vocab_from_word:
    vocab_from_word_in_list.append(i)




# import the vocab list and put it into a list
# it's tuples... but i think it will be OK

wb = openpyxl.load_workbook('file1.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
vocab_list = []
for i in range(1, 114, 1):
#    print (i, sheet.cell(row=i, column=2).value)
   character = i, sheet.cell(row=i, column=1).value
   vocab_list.append(character)
print (vocab_list)


# get some basic stats.

characters_in_story = len(story_list)
print ('there are ', characters_in_story, 'characters in the story')

characters_in_vocab_list = len(vocab_from_word_in_list)
print ('there are ', characters_in_vocab_list, 'characters in the vocab list')



########### EXCEL FILE PARKED HERE FOR A BIT  ####################################
### characters_in_vocab_list = len(vocab_list)
### print ('there are ', characters_in_vocab_list, 'characters in the vocab list') 
##################################################################################


# work out percentage of characters on the vocab list that were used in the story.

print(set(story_list)) # set() removes duplicates

count = 0
characters_used = []
for i in set(story_list):
  if i in vocab_from_word_in_list:
    print ('the vocab_list character used in the story is :' , i)
    count += 1
    characters_used.append(i)

print ('count of characters used in the story is ', count)
print ('the list of the characters that you used is:', characters_used)



print ('count of potential characters in the vocab list which you could have used are:')

print ('you used ', int(count/characters_in_vocab_list*100), '% of possible vocab')


# show the characters I did not use, and therefore should use:

count = 0
unused_characters = []

unused_characters = set(vocab_from_word_in_list) - set(characters_used)

# for i in set(story_list):
#     x = i for i in vocab_from_word_in_list if i not in characters_used:
#     # if i in vocab_from_word_in_list and i not in characters_used:
#         unused_characters.append(x)
#         print (i)


print ('your unused characters are ', unused_characters)
print  ('you didn\'t use', len(unused_characters), 'characters') 

print ("you didn't use", len(unused_characters), 'characters')